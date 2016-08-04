# python3.
import openpyxl
from openpyxl.styles import Font, Style
import os
# that takes a number N from the command
# line and creates an N×N multiplication table in an Excel spreadsheet.

os.chdir(r'C:\Users\Wambayi\Desktop\desktopfolders\Python\Bins\Tinkering')


def multplicationTable(n):
    wb = openpyxl.Workbook()
    multTable = wb.active
    myFont = Font(bold=True)
    myStyle = Style(font=myFont)
    for i in range(2, n + 2):
        multTable.cell(row=1, column=i, value=i - 1)
        multTable.cell(row=i, column=1, value=i - 1)
    for row in range(2, n + 2):
        for col in range(2, n + 2):
            multTable.cell(row=row, column=col).value = multTable.cell(
                row=1, column=col).value * multTable.cell(row=row, column=1).value
    wb.save('multTable.xlsx')
if __name__ == '__main__':
    multplicationTable(6)
